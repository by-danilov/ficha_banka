import json
import csv
import logging
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Настройка логирования для file_operations.py
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

log_dir = 'logs'
os.makedirs(log_dir, exist_ok=True)
file_handler = logging.FileHandler(os.path.join(log_dir, 'file_operations.log'), mode='w', encoding='utf-8')
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


def load_operations_from_json(json_filepath: str) -> list[dict]:
    """
    Загружает список финансовых операций из JSON файла.
    Возвращает пустой список, если файл не найден, пуст или содержит некорректные данные.
    """
    if not os.path.exists(json_filepath):
        logger.error(f"JSON файл не найден: {json_filepath}")
        return []

    try:
        with open(json_filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if not isinstance(data, list):
                logger.error(f"JSON файл {json_filepath} содержит некорректный формат данных (ожидается список).")
                return []
            # Фильтруем пустые или некорректные записи
            operations = [op for op in data if isinstance(op, dict) and op.get('id') is not None]
            logger.info(f"Успешно загружено {len(operations)} операций из JSON файла: {json_filepath}")
            return operations
    except json.JSONDecodeError as e:
        logger.error(f"Ошибка декодирования JSON файла {json_filepath}: {e}")
        return []
    except Exception as e:
        logger.error(f"Неожиданная ошибка при чтении JSON файла {json_filepath}: {e}")
        return []


def read_operations_from_csv(csv_filepath: str) -> list[dict]:
    """
    Читает список финансовых операций из CSV файла.
    Ожидает разделитель ';'.
    """
    if not os.path.exists(csv_filepath):
        logger.error(f"CSV файл не найден: {csv_filepath}")
        return []

    operations = []
    # Обновленные обязательные заголовки для CSV с учетом currency_code
    required_headers = ['id', 'state', 'date', 'amount', 'currency_name', 'currency_code', 'description']

    try:
        with open(csv_filepath, 'r', encoding='utf-8') as file:
            # Изменяем разделитель на ';'
            reader = csv.DictReader(file, delimiter=';')

            if not all(header in reader.fieldnames for header in required_headers):
                missing_headers = [header for header in required_headers if header not in reader.fieldnames]
                logger.error(f"CSV файл '{csv_filepath}' не содержит всех обязательных заголовков: {missing_headers}. "
                             f"Найдено: {reader.fieldnames}")
                return []

            for i, row in enumerate(reader):
                # Пропускаем полностью пустые строки (например, если в конце файла есть лишняя пустая строка)
                if not any(row.values()):
                    continue

                try:
                    # Приводим 'state' к верхнему регистру для единообразия с JSON
                    status = str(row.get('state', '')).upper()

                    operation = {
                        'id': int(row['id']),
                        'description': str(row['description']),
                        'amount': float(row['amount']),
                        'currency_name': str(row['currency_name']),  # Добавлено
                        'currency_code': str(row['currency_code']),  # Добавлено, ключевой для фильтрации
                        'date': str(row['date']),  # Дата остается строкой, форматируется позже
                        'state': status,  # Используем 'state' вместо 'status'
                        'from': str(row.get('from', '')),
                        'to': str(row.get('to', ''))
                    }
                    operations.append(operation)
                except KeyError as e:
                    logger.error(
                        f"Ошибка при обработке строки {i + 1} в CSV файле '{csv_filepath}': Отсутствует ключ {e}. Строка: {row}")
                    continue
                except ValueError as e:
                    logger.error(
                        f"Ошибка преобразования данных в строке {i + 1} в CSV файле '{csv_filepath}': {e}. Строка: {row}")
                    continue
                except Exception as e:
                    logger.error(
                        f"Неожиданная ошибка при обработке строки {i + 1} в CSV файле '{csv_filepath}': {e}. Строка: {row}")
                    continue
        logger.info(f"Успешно загружено {len(operations)} операций из CSV файла: {csv_filepath}")
        return operations
    except Exception as e:
        logger.error(f"Ошибка при чтении CSV файла {csv_filepath}: {e}")
        return []


def read_operations_from_excel(excel_filepath: str) -> list[dict]:
    """
    Читает список финансовых операций из Excel файла.
    """
    if not os.path.exists(excel_filepath):
        logger.error(f"Excel файл не найден: {excel_filepath}")
        return []

    operations = []
    # Обновленные обязательные заголовки для Excel с учетом currency_code
    required_headers = ['id', 'state', 'date', 'amount', 'currency_name', 'currency_code', 'description']

    try:
        workbook = load_workbook(excel_filepath)
        sheet = workbook.active

        # Получаем заголовки из первой строки
        header = [cell.value for cell in sheet[1]]

        if not all(h in header for h in required_headers):
            missing_headers = [h for h in required_headers if h not in header]
            logger.error(f"Excel файл '{excel_filepath}' не содержит всех обязательных заголовков: {missing_headers}. "
                         f"Найдено: {header}")
            return []

        # Создаем маппинг колонок по именам заголовков
        header_map = {h: col_idx for col_idx, h in enumerate(header)}

        for row_index in range(2, sheet.max_row + 1):  # Начинаем со второй строки (после заголовков)
            row_data = {}
            for h in header:
                col_idx = header_map[h]
                cell_value = sheet.cell(row=row_index, column=col_idx + 1).value
                row_data[h] = cell_value

            # Пропускаем полностью пустые строки
            if not any(value for key, value in row_data.items() if
                       key not in ['from', 'to', 'description'] and value is not None):
                continue

            try:
                date_value = row_data.get('date')
                if isinstance(date_value, datetime):
                    date_str = date_value.strftime('%Y-%m-%dT%H:%M:%SZ')  # Сохраняем формат JSON
                elif date_value is not None:
                    date_str = str(date_value)
                else:
                    date_str = ''

                # Приводим 'state' к верхнему регистру для единообразия с JSON
                status = str(row_data.get('state', '')).upper()

                operation = {
                    'id': int(row_data['id']),
                    'description': str(row_data['description']),
                    'amount': float(row_data['amount']),
                    'currency_name': str(row_data['currency_name']),  # Добавлено
                    'currency_code': str(row_data['currency_code']),  # Добавлено, ключевой для фильтрации
                    'date': date_str,
                    'state': status,
                    'from': str(row_data.get('from', '') if row_data.get('from') is not None else ''),
                    'to': str(row_data.get('to', '') if row_data.get('to') is not None else '')
                }
                operations.append(operation)
            except KeyError as e:
                logger.error(
                    f"Ошибка при обработке строки {row_index} в Excel файле '{excel_filepath}': Отсутствует ключ {e}. Строка: {row_data}")
                continue
            except ValueError as e:
                logger.error(
                    f"Ошибка преобразования данных в строке {row_index} в Excel файле '{excel_filepath}': {e}. Строка: {row_data}")
                continue
            except Exception as e:
                logger.error(
                    f"Неожиданная ошибка при обработке строки {row_index} в Excel файле '{excel_filepath}': {e}. Строка: {row_data}")
                continue
        logger.info(f"Успешно загружено {len(operations)} операций из Excel файла: {excel_filepath}")
        return operations
    except InvalidFileException as e:
        logger.error(f"Ошибка: Файл '{excel_filepath}' не является действительным файлом Excel. {e}")
        return []
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла {excel_filepath}: {e}")
        return []
