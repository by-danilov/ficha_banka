import os
from datetime import datetime
from unittest.mock import MagicMock, mock_open, patch

import pytest

# Обратите внимание, что здесь импорт из file_operations/file_operations.py
from src.file_operations.file_operations import (
    read_operations_from_csv,
    read_operations_from_excel,
)

# --- Тесты для read_operations_from_csv ---


def test_read_operations_from_csv_success_mock():
    csv_data = (
        "id;state;date;amount;currency_name;currency_code;description;from;to\n"
        "1;EXECUTED;2023-01-15;100.50;Рубли;RUB;Оплата;Карта 1;Счет 1\n"
        "2;PENDING;2023-01-16;200.00;Доллары;USD;Перевод;Счет 2;Карта 2\n"
        "3;CANCELED;2023-01-17;50.25;Евро;EUR;Покупка;;Счет 3\n"  # Пустое "from"
    )
    mock_file = mock_open(read_data=csv_data)

    # Здесь os.path.exists мокается, так как функция CSV его использует
    with (
        patch("builtins.open", mock_file),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_csv("dummy_path.csv")

        expected_operations = [
            {
                "id": 1,
                "description": "Оплата",
                "amount": 100.50,
                "currency_name": "Рубли",
                "currency_code": "RUB",
                "date": "2023-01-15",
                "state": "EXECUTED",
                "from": "Карта 1",
                "to": "Счет 1",
            },
            {
                "id": 2,
                "description": "Перевод",
                "amount": 200.00,
                "currency_name": "Доллары",
                "currency_code": "USD",
                "date": "2023-01-16",
                "state": "PENDING",
                "from": "Счет 2",
                "to": "Карта 2",
            },
            {
                "id": 3,
                "description": "Покупка",
                "amount": 50.25,
                "currency_name": "Евро",
                "currency_code": "EUR",
                "date": "2023-01-17",
                "state": "CANCELED",
                "from": "",  # Ожидаем пустую строку для отсутствующего значения
                "to": "Счет 3",
            },
        ]
        assert operations == expected_operations
        mock_file.assert_called_once_with("dummy_path.csv", "r", encoding="utf-8")


def test_read_operations_from_csv_empty_file_mock():
    mock_file = mock_open(read_data="")
    with (
        patch("builtins.open", mock_file),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_csv("dummy_path.csv")
        assert operations == []


def test_read_operations_from_csv_no_header_mock():
    csv_data = "value1;value2\nvalue3;value4\n"
    mock_file = mock_open(read_data=csv_data)
    with (
        patch("builtins.open", mock_file),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_csv("dummy_path.csv")
        assert operations == []


def test_read_operations_from_csv_missing_required_headers():
    csv_data = (
        "id;state;date;amount;currency_name;from;to\n"
        "1;EXECUTED;2023-01-15;100.50;Рубли;Карта 1;Счет 1\n"
    )
    mock_file = mock_open(read_data=csv_data)
    with (
        patch("builtins.open", mock_file),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_csv("dummy_path.csv")
        assert operations == []


def test_read_operations_from_csv_invalid_data_types():
    csv_data = (
        "id;state;date;amount;currency_name;currency_code;description;from;to\n"
        "abc;EXECUTED;2023-01-15;not_a_number;Рубли;RUB;Оплата;Карта 1;Счет 1\n"  # Некорректный id и amount
    )
    mock_file = mock_open(read_data=csv_data)
    with (
        patch("builtins.open", mock_file),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_csv("dummy_path.csv")
        assert operations == []  # Ожидаем пустой список, т.к. строка будет пропущена


def test_read_operations_from_csv_file_not_found_mock():
    with patch(
        "src.file_operations.file_operations.os.path.exists", return_value=False
    ):
        operations = read_operations_from_csv("non_existent.csv")
        assert operations == []


# --- Тесты для read_operations_from_excel ---


# Вспомогательная функция для создания mock-ячейки
def create_mock_cell(value):
    mock_cell = MagicMock()
    mock_cell.value = value
    return mock_cell


def test_read_operations_from_excel_success_mock():
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.active = mock_sheet

    headers = [
        "id",
        "state",
        "date",
        "amount",
        "currency_name",
        "currency_code",
        "description",
        "from",
        "to",
    ]

    mock_sheet.__getitem__.return_value = [create_mock_cell(h) for h in headers]

    mock_sheet.max_row = 3  # Заголовки + 2 строки данных
    mock_sheet.max_column = len(headers)

    cell_values = {
        (1, 1): "id",
        (1, 2): "state",
        (1, 3): "date",
        (1, 4): "amount",
        (1, 5): "currency_name",
        (1, 6): "currency_code",
        (1, 7): "description",
        (1, 8): "from",
        (1, 9): "to",
        (2, 1): 1,
        (2, 2): "EXECUTED",
        (2, 3): datetime(2023, 1, 15, 12, 0, 0),
        (2, 4): 100.50,
        (2, 5): "Рубли",
        (2, 6): "RUB",
        (2, 7): "Оплата",
        (2, 8): "Карта 1",
        (2, 9): "Счет 1",
        (3, 1): 2,
        (3, 2): "PENDING",
        (3, 3): datetime(2023, 1, 16, 9, 30, 0),
        (3, 4): 200.00,
        (3, 5): "Доллары",
        (3, 6): "USD",
        (3, 7): "Перевод",
        (3, 8): "Счет 2",
        (3, 9): "Карта 2",
    }

    def mock_cell_access(row, column):
        return create_mock_cell(cell_values.get((row, column)))

    mock_sheet.cell.side_effect = mock_cell_access

    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути, где он используется
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути, где он используется
    with (
        patch(
            "src.file_operations.file_operations.load_workbook",
            return_value=mock_workbook,
        ),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_excel("dummy_path.xlsx")

        expected_operations = [
            {
                "id": 1,
                "description": "Оплата",
                "amount": 100.50,
                "currency_name": "Рубли",
                "currency_code": "RUB",
                "date": "2023-01-15T12:00:00Z",  # Excel-функция сохраняет Z для ISO
                "state": "EXECUTED",
                "from": "Карта 1",
                "to": "Счет 1",
            },
            {
                "id": 2,
                "description": "Перевод",
                "amount": 200.00,
                "currency_name": "Доллары",
                "currency_code": "USD",
                "date": "2023-01-16T09:30:00Z",  # Excel-функция сохраняет Z для ISO
                "state": "PENDING",
                "from": "Счет 2",
                "to": "Карта 2",
            },
        ]
        assert operations == expected_operations


def test_read_operations_from_excel_empty_file_mock():
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.active = mock_sheet

    headers = [
        "id",
        "state",
        "date",
        "amount",
        "currency_name",
        "currency_code",
        "description",
        "from",
        "to",
    ]
    mock_sheet.__getitem__.return_value = [create_mock_cell(h) for h in headers]

    mock_sheet.max_row = 1  # Только заголовки
    mock_sheet.max_column = len(headers)
    mock_sheet.cell.side_effect = lambda row, column: create_mock_cell(
        None
    )  # Все ячейки пустые

    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути, где он используется
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути, где он используется
    with (
        patch(
            "src.file_operations.file_operations.load_workbook",
            return_value=mock_workbook,
        ),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_excel("dummy_path.xlsx")
        assert operations == []


def test_read_operations_from_excel_no_header_mock():
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.active = mock_sheet

    # Возвращаем некорректные заголовки
    mock_sheet.__getitem__.return_value = [
        create_mock_cell("value1"),
        create_mock_cell("value2"),
        create_mock_cell("value3"),
    ]
    mock_sheet.max_row = 1
    mock_sheet.max_column = 3
    mock_sheet.cell.side_effect = lambda row, column: create_mock_cell(None)

    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути, где он используется
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути, где он используется
    with (
        patch(
            "src.file_operations.file_operations.load_workbook",
            return_value=mock_workbook,
        ),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_excel("dummy_path.xlsx")
        assert operations == []


def test_read_operations_from_excel_missing_required_headers():
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.active = mock_sheet

    headers = [
        "id",
        "state",
        "date",
        "amount",
        "description",
        "from",
        "to",
    ]  # Отсутствуют currency_name, currency_code
    mock_sheet.__getitem__.return_value = [create_mock_cell(h) for h in headers]

    mock_sheet.max_row = 2
    mock_sheet.max_column = len(headers)

    cell_values = {
        (1, 1): "id",
        (1, 2): "state",
        (1, 3): "date",
        (1, 4): "amount",
        (1, 5): "description",
        (1, 6): "from",
        (1, 7): "to",
        (2, 1): 1,
        (2, 2): "EXECUTED",
        (2, 3): datetime(2023, 1, 15),
        (2, 4): 100.50,
        (2, 5): "Оплата",
        (2, 6): "Карта 1",
        (2, 7): "Счет 1",
    }
    mock_sheet.cell.side_effect = lambda row, column: create_mock_cell(
        cell_values.get((row, column))
    )

    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути, где он используется
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути, где он используется
    with (
        patch(
            "src.file_operations.file_operations.load_workbook",
            return_value=mock_workbook,
        ),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_excel("dummy_path.xlsx")
        assert operations == []


def test_read_operations_from_excel_invalid_data_types():
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.active = mock_sheet

    headers = [
        "id",
        "state",
        "date",
        "amount",
        "currency_name",
        "currency_code",
        "description",
        "from",
        "to",
    ]
    mock_sheet.__getitem__.return_value = [create_mock_cell(h) for h in headers]

    mock_sheet.max_row = 2
    mock_sheet.max_column = len(headers)

    cell_values = {
        (1, 1): "id",
        (1, 2): "state",
        (1, 3): "date",
        (1, 4): "amount",
        (1, 5): "currency_name",
        (1, 6): "currency_code",
        (1, 7): "description",
        (1, 8): "from",
        (1, 9): "to",
        (2, 1): 1,
        (2, 2): "EXECUTED",
        (2, 3): datetime(2023, 1, 15),
        (2, 4): "not_a_number",
        # amount - некорректный тип
        (2, 5): "Рубли",
        (2, 6): "RUB",
        (2, 7): "Оплата",
        (2, 8): "Карта 1",
        (2, 9): "Счет 1",
    }
    mock_sheet.cell.side_effect = lambda row, column: create_mock_cell(
        cell_values.get((row, column))
    )

    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути, где он используется
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути, где он используется
    with (
        patch(
            "src.file_operations.file_operations.load_workbook",
            return_value=mock_workbook,
        ),
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
    ):
        operations = read_operations_from_excel("dummy_path.xlsx")
        assert operations == []


def test_read_operations_from_excel_file_not_found_mock():
    # Здесь os.path.exists по-прежнему должен быть False, чтобы функция вернула []
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути
    with patch(
        "src.file_operations.file_operations.os.path.exists", return_value=False
    ):
        operations = read_operations_from_excel("non_existent.xlsx")
        assert operations == []


def test_read_operations_from_excel_invalid_file_exception():
    """Тест, когда load_workbook вызывает InvalidFileException."""
    from openpyxl.utils.exceptions import InvalidFileException

    # Здесь os.path.exists должен быть True, иначе тест будет пропущен раньше
    # ИСПРАВЛЕНО: Мокируем load_workbook по полному пути
    # ИСПРАВЛЕНО: Мокируем os.path.exists по полному пути
    with (
        patch("src.file_operations.file_operations.os.path.exists", return_value=True),
        patch(
            "src.file_operations.file_operations.load_workbook",
            side_effect=InvalidFileException("Invalid file format"),
        ),
    ):
        operations = read_operations_from_excel("invalid.xlsx")
        assert operations == []
